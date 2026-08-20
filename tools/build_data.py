"""Precompute everything the browser would otherwise derive on every page load.

Reads data/properties.xlsx + the boundary files, resolves each property's
submarket / city / county by point-in-polygon, and writes data/derived.json.
The app loads that file when present (fast path) and falls back to parsing the
workbook in-browser when it is absent or stale.

Run after any data change:  python3 tools/build_data.py
"""
import json, pathlib, zipfile, datetime
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from shapely.geometry import Point, shape
from shapely.strtree import STRtree

ROOT = pathlib.Path(__file__).resolve().parent.parent
OUT = ROOT / "data" / "derived.json"
KML_NS = "{http://www.opengis.net/kml/2.2}"


def load_config():
    return json.loads((ROOT / "config.json").read_text())


def parse_kml_polygons(text):
    """-> [(name, shapely geometry)] for every Placemark polygon in a KML document."""
    root = ET.fromstring(text)
    out = []
    for pm in root.iter(KML_NS + "Placemark"):
        name_el = pm.find(KML_NS + "name")
        name = (name_el.text or "").strip() if name_el is not None else ""
        polys = []
        for poly in pm.iter(KML_NS + "Polygon"):
            outer = poly.find(f"{KML_NS}outerBoundaryIs/{KML_NS}LinearRing/{KML_NS}coordinates")
            if outer is None or not (outer.text or "").strip():
                continue
            ring = [tuple(float(x) for x in c.split(",")[:2]) for c in outer.text.split()]
            holes = []
            for inner in poly.findall(f"{KML_NS}innerBoundaryIs/{KML_NS}LinearRing/{KML_NS}coordinates"):
                if (inner.text or "").strip():
                    holes.append([tuple(float(x) for x in c.split(",")[:2]) for c in inner.text.split()])
            if len(ring) >= 4:
                polys.append(shape({"type": "Polygon", "coordinates": [ring] + holes}))
        for g in polys:
            out.append((name, g))
    return out


def load_boundaries(paths):
    feats = []
    for rel in paths:
        p = ROOT / rel
        if not p.exists():
            print(f"  ! missing {rel}")
            continue
        if p.suffix.lower() in (".kmz", ".zip"):
            with zipfile.ZipFile(p) as z:
                kml = next((n for n in z.namelist() if n.lower().endswith(".kml")), None)
                if not kml:
                    continue
                feats += parse_kml_polygons(z.read(kml).decode("utf-8", "replace"))
        elif p.suffix.lower() == ".kml":
            feats += parse_kml_polygons(p.read_text())
        elif p.suffix.lower() in (".geojson", ".json"):
            gj = json.loads(p.read_text())
            for f in gj.get("features", []):
                props = f.get("properties") or {}
                name = props.get("name") or props.get("Name") or props.get("NAME") or ""
                if f.get("geometry"):
                    feats.append((str(name).strip(), shape(f["geometry"])))
    return feats


class Locator:
    """Point-in-polygon index over (name, geometry) pairs."""

    def __init__(self, feats):
        self.names = [n for n, _ in feats]
        self.geoms = [g for _, g in feats]
        self.tree = STRtree(self.geoms) if self.geoms else None

    def find(self, lng, lat):
        if not self.tree:
            return ""
        pt = Point(lng, lat)
        for i in self.tree.query(pt):
            if self.geoms[i].contains(pt):
                return self.names[i]
        return ""


def main():
    cfg = load_config()
    f = cfg["fields"]
    subs_cfg = cfg["layers"]["submarkets"]
    sub_paths = subs_cfg if isinstance(subs_cfg, list) else subs_cfg.get("default", [])
    admin = cfg["layers"].get("admin", {})

    print("loading boundaries…")
    submarkets = Locator(load_boundaries(sub_paths))
    cities = Locator(load_boundaries(admin.get("cities", {}).get("files", [])))
    counties = Locator(load_boundaries(admin.get("counties", {}).get("files", [])))

    wb = load_workbook(ROOT / cfg["data"]["workbook"], read_only=True, data_only=True)
    ws = wb[cfg["data"].get("sheet")] if cfg["data"].get("sheet") in wb.sheetnames else wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h) if h is not None else "" for h in next(rows_iter)]

    out_rows, resolved = [], {"submarket": 0, "city": 0, "county": 0}
    for raw in rows_iter:
        row = {h: (v if v is not None else "") for h, v in zip(headers, raw) if h}
        if not row.get(f["name"]) and not row.get(f["id"]):
            continue
        try:
            lat, lng = float(row.get(f["lat"])), float(row.get(f["lng"]))
        except (TypeError, ValueError):
            continue
        # workbook values always win; geometry fills the blanks
        if not str(row.get(f["submarket"], "")).strip():
            hit = submarkets.find(lng, lat)
            if hit:
                row[f["submarket"]] = hit
                resolved["submarket"] += 1
        if not str(row.get(f["city"], "")).strip():
            hit = cities.find(lng, lat)
            if hit:
                row[f["city"]] = hit
                resolved["city"] += 1
        if not str(row.get(f["county"], "")).strip():
            hit = counties.find(lng, lat)
            if hit:
                row[f["county"]] = hit.replace(" County", "")
                resolved["county"] += 1
        # dates serialize as ISO strings so the browser reads them uniformly
        for k, v in list(row.items()):
            if isinstance(v, (datetime.datetime, datetime.date)):
                row[k] = v.strftime("%Y-%m-%d")
        out_rows.append(row)

    payload = {
        "builtAt": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        "source": cfg["data"]["workbook"],
        "columns": [h for h in headers if h],
        "rows": out_rows,
    }
    OUT.write_text(json.dumps(payload, separators=(",", ":")))
    kb = OUT.stat().st_size / 1024
    print(f"wrote {len(out_rows)} rows x {len(payload['columns'])} cols -> {OUT.name} ({kb:.0f} KB)")
    print(f"  resolved by geometry: {resolved}")


if __name__ == "__main__":
    main()
