"""Fetch REAL parcel boundaries for every property in data/properties.xlsx from free
county open-data ArcGIS services (Hennepin + Ramsey), writing data/parcels/msp_parcels.geojson.

Re-run whenever properties change:  python3 tools/fetch_parcels.py
Add more counties by extending SERVICES (any ArcGIS FeatureServer/MapServer polygon layer works).
"""
import json, time, urllib.parse, urllib.request, pathlib

ROOT = pathlib.Path(__file__).resolve().parent.parent
OUT = ROOT / "data" / "parcels" / "msp_parcels.geojson"

SERVICES = {
    "Hennepin": "https://gis.hennepin.us/arcgis/rest/services/HennepinData/LAND_PROPERTY/MapServer/1/query",
    "Ramsey": "https://maps.co.ramsey.mn.us/arcgis/rest/services/ParcelData/AttributedData/MapServer/2/query",
}

def run_query(url, params):
    req = urllib.request.Request(url + "?" + urllib.parse.urlencode(params), headers={"User-Agent": "idmt-parcel-fetch"})
    with urllib.request.urlopen(req, timeout=30) as r:
        data = json.load(r)
    return data.get("features", [])

def query_parcel(url, lng, lat):
    base = {
        "geometryType": "esriGeometryPoint",
        "inSR": "4326",
        "spatialRel": "esriSpatialRelIntersects",
        "outFields": "",
        "returnGeometry": "true",
        "outSR": "4326",
        "f": "geojson",
    }
    feats = run_query(url, dict(base, geometry=f"{lng},{lat}"))
    if not feats:
        # point landed on a street/right-of-way: search a ~50m envelope and take the nearest hit
        d = 0.00045
        feats = run_query(url, dict(base,
            geometry=f"{lng - d},{lat - d},{lng + d},{lat + d}",
            geometryType="esriGeometryEnvelope"))
    return feats[0] if feats else None

def main():
    from openpyxl import load_workbook
    wb = load_workbook(ROOT / "data" / "properties.xlsx", read_only=True)
    ws = wb["Properties"]
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    features, misses = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, county = row[idx["Property Name"]], row[idx["County"]]
        lat, lng = row[idx["Latitude"]], row[idx["Longitude"]]
        if not (lat and lng):
            continue
        url = SERVICES.get(str(county).strip())
        if not url:
            misses.append((name, county, "no service for county"))
            continue
        try:
            feat = query_parcel(url, lng, lat)
        except Exception as e:
            misses.append((name, county, str(e)[:80]))
            continue
        if feat:
            feat["properties"] = {"name": f"{name} parcel", "county": county}
            features.append(feat)
        else:
            misses.append((name, county, "no parcel at point"))
        time.sleep(0.25)  # be polite to county servers
    OUT.write_text(json.dumps({"type": "FeatureCollection", "features": features}))
    print(f"wrote {len(features)} parcels -> {OUT}")
    for m in misses:
        print("  miss:", m)

if __name__ == "__main__":
    main()
